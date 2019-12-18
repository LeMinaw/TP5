import openpyxl
import unicodedata
from bs4 import BeautifulSoup
from urllib.request import urlopen


def remove_accents(string):
    """Shorthand function for making strings ASCII-compatible."""

    nkfd_form = unicodedata.normalize('NFKD', string)
    return ''.join([c for c in nkfd_form if not unicodedata.combining(c)])


def get_card_data(card_html):
    """Gets data from provided <card_html> HTML content."""
    
    country = card_html.select_one('.lvdetails').select('li')
    if len(country) > 1:
        country = country[1].text.replace("Provenance\xa0: ", '').strip()
    else:
        country = "France"
    
    return {
        'name': card_html.select_one('h3 a').text,
        'price': card_html.select_one('.prc .bidsold').text.split()[0],
        'date': card_html.select_one('.lvdetails .tme span').text,
        'country': country,
    }


def get_card_html(card_name):
    """Performs a search of <card_name> and returns a list of HTML content for each result."""

    card_format = remove_accents(card_name).replace(' ', '%20')
    html = urlopen(f"https://www.ebay.fr/sch/i.html?_from=R40&_sacat=0&_nkw={card_format}&LH_Complete=1&LH_Sold=1&rt=ncRemarquer")
    soup = BeautifulSoup(html, 'html.parser')
    return soup.find_all("li", _sp="p2045573.m1686.l0", attrs={"class": "sresult"})


def browse_card_file(file_path):
    """Browses a source file and returns data about the card it contains."""

    wb = openpyxl.load_workbook(file_path)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    return [{
        'name': row[1],
        'id': row[2]
    } for row in ws.iter_rows(min_row=3, values_only=True)]


def build_cards_data_file(cards_file_path, result_file_path):
    """Builds dataset by performing requests on cards in given cards file."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("Nom", "Identifiant", "Nom de l'annonce",
        "Prix de vente", "Date de vente", "Provenance"))
    
    for card in browse_card_file(cards_file_path):
        card_query = f"{card['name']} {card['id']}"
        for card_html in get_card_html(card_query):
            card_data = get_card_data(card_html)
            ws.append((
                card['name'],
                card['id'],
                card_data['name'],
                card_data['price'],
                card_data['date'],
                card_data['country']
            ))
        print(f"Done card {card['name']}.")
    wb.save(filename=result_file_path)


if __name__ == '__main__':
    # for card_html in get_card_html("Pikachu 48/102"):
    #     print(get_card_data(card_html))
    # for card in browse_card_file("cartes.xlsx"):
    #     print(card)
    build_cards_data_file("cartes.xlsx", "données_cartes.xlsx")
